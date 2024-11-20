import logging
import os

import pandas as pd
import sys
import json

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# from utils import show_message

logging.basicConfig(level=logging.INFO)

def load_excel(file_path):
    return pd.read_excel(file_path)


def calculate_forecast(df, references_month, references_year, target_start_year, target_end_year, specifics_enabled,
                       prod_nums, bus_chanl_nums):
    print("Filtering reference data based on provided month and year...")
    reference_data_current_year = df[
        (df['PERIOD_YEAR'] == references_year) &
        (df['PERIOD_MONTH'] <= references_month)
        ]
    reference_data_previous_year = df[
        (df['PERIOD_YEAR'] == (references_year - 1)) &
        (df['PERIOD_MONTH'] > references_month)
        ]
    reference_data = pd.concat([reference_data_previous_year, reference_data_current_year])
    print(f"Reference data after initial filter: {len(reference_data)} rows")

    if specifics_enabled:
        print("Filtering reference data based on specifics...")
        print(f"Selected PROD_NUMs: {prod_nums}")
        print(f"Selected BUS_CHANL_NUMs: {bus_chanl_nums}")
        unique_prod_nums = reference_data['PROD_NUM'].astype(str).unique()
        unique_bus_chanl_nums = reference_data['BUS_CHANL_NUM'].astype(str).unique()
        print(f"Unique PROD_NUMs in reference data: {unique_prod_nums}")
        print(f"Unique BUS_CHANL_NUMs in reference data: {unique_bus_chanl_nums}")

        if not prod_nums:
            prod_nums = unique_prod_nums.tolist()
        if not bus_chanl_nums:
            bus_chanl_nums = unique_bus_chanl_nums.tolist()

        reference_data = reference_data[
            (reference_data['PROD_NUM'].astype(str).isin(prod_nums)) &
            (reference_data['BUS_CHANL_NUM'].astype(str).isin(bus_chanl_nums))
            ]
        print(f"Reference data after specifics filter: {len(reference_data)} rows")

    print("Checking for duplicates...")
    duplicates = reference_data[
        reference_data.duplicated(subset=['PERIOD_YEAR', 'PERIOD_MONTH', 'PROD_NUM', 'BUS_CHANL_NUM'], keep=False)
    ]
    if not duplicates.empty:
        print("Duplicates found, generating error message...")
        duplicate_info = duplicates[['PERIOD_YEAR', 'PERIOD_MONTH', 'PROD_NUM', 'BUS_CHANL_NUM']].drop_duplicates()
        duplicate_details = "\n".join([
            f"Year: {row.PERIOD_YEAR}, Month: {row.PERIOD_MONTH}, Prod Num: {row.PROD_NUM}, Bus Chanl Num: {row.BUS_CHANL_NUM}"
            for idx, row in duplicate_info.iterrows()
        ])

        duplicate_rows = duplicates.index.tolist()
        duplicate_rows_info = "\n".join([f"Row Number: {row_num}" for row_num in duplicate_rows])

        error_message = f"Duplicate rows found in the reference file based on 'PERIOD_YEAR', 'PERIOD_MONTH', 'PROD_NUM', 'BUS_CHANL_NUM':\n{duplicate_details}\n\nDuplicate Rows:\n{duplicate_rows_info}"
        # show_message("Error", error_message, type='error')
        return pd.DataFrame(), pd.DataFrame()

    print("Calculating reference eop volumes...")
    eop_2024 = reference_data.groupby(['PERIOD_YEAR', 'PERIOD_MONTH', 'PROD_NUM', 'BUS_CHANL_NUM'])[
        'sum_eop_vol_2024'].sum()
    eop_2025 = reference_data.groupby(['PERIOD_YEAR', 'PERIOD_MONTH', 'PROD_NUM', 'BUS_CHANL_NUM'])[
        'sum_eop_vol_2025'].sum()

    forecast_data = []

    print("Starting forecast calculation...")
    for year in range(target_start_year, target_end_year + 1):
        for month in range(1, 13):
            if month <= references_month:
                ref_period_year = references_year
                ref_period_month = month
            else:
                ref_period_year = references_year - 1
                ref_period_month = month

            ref_data = reference_data[
                (reference_data['PERIOD_YEAR'] == ref_period_year) &
                (reference_data['PERIOD_MONTH'] == ref_period_month)
                ]

            for index, row in ref_data.iterrows():
                prod_num = row['PROD_NUM']
                bus_chanl_num = row['BUS_CHANL_NUM']

                eop_2024_val = eop_2024.get((ref_period_year, ref_period_month, prod_num, bus_chanl_num), float('nan'))
                eop_2025_val = eop_2025.get((ref_period_year, ref_period_month, prod_num, bus_chanl_num), float('nan'))

                forecast_row = row.copy()
                if not pd.isna(eop_2024_val) and eop_2024_val != 0 and not pd.isna(eop_2025_val):
                    for col in ['LIVE_TV_VIEWING_MINUTES', 'PVR_VIEWING_MINUTES', 'CUTV_VIEWING_MINUTES',
                                'OTT_VIEWING_MINUTES', 'VOD_VIEWING_MINUTES']:
                        forecasted_viewing = row[col] * eop_2025_val / eop_2024_val
                        forecast_row[col] = forecasted_viewing
                forecast_row['PERIOD_YEAR'] = year
                forecast_row['PERIOD_MONTH'] = month
                forecast_data.append(forecast_row.to_dict())

    print(f"Forecast calculation completed. Total forecast rows: {len(forecast_data)}")
    return pd.DataFrame(forecast_data), reference_data



def copy_sheet(source_sheet, target_sheet):
    for row in source_sheet.iter_rows():
        for cell in row:
            target_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)

    for row_dim in source_sheet.row_dimensions.values():
        target_sheet.row_dimensions[row_dim.index].height = row_dim.height

    for col_dim in source_sheet.column_dimensions.values():
        target_sheet.column_dimensions[col_dim.index].width = col_dim.width

    for merged_cell_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_cell_range))

def check_file_open(file_path):
    if not os.path.isfile(file_path):
        return False
    try:
        os.rename(file_path, file_path)
    except OSError:
        return True
    return False

def style_worksheet(ws):
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill(start_color="4ea72e", end_color="4ea72e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alternating_fill = [PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
                        PatternFill(start_color="daf2d0", end_color="daf2d0", fill_type="solid")]
    border = Border(top=Side(style="thin", color="4ea72e"), bottom=Side(style="thin", color="4ea72e"))

    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.fill = alternating_fill[(cell.row - 2) % 2]

    for col in ws.iter_cols(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2 if max_length > 0 else 8
        ws.column_dimensions[column_letter].width = adjusted_width

def save_dataframe_with_formatting(forecast_df, reference_df, output_path, original_file, references_year, prod_nums, bus_chanl_nums):
    if not os.path.exists(output_path):
        os.makedirs(output_path)

    output_filepath = os.path.join(output_path, "forecast_audience.xlsx")

    if check_file_open(output_filepath):
        logging.error(f"The file {output_filepath} is open. Please close the file and try again.")
        return

    try:
        logging.info(f"Loading original workbook from {original_file}")
        workbook = load_workbook(original_file)
        reference_sheet = workbook.active

        forecast_sheet = workbook.create_sheet(title="Working")
        new_reference_sheet = workbook.create_sheet(title="Reference")

        logging.info("Writing data to the Working sheet")
        for r_idx, row in enumerate(dataframe_to_rows(forecast_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                forecast_sheet.cell(row=r_idx, column=c_idx, value=value)

        logging.info("Writing data to the Reference sheet")
        for r_idx, row in enumerate(dataframe_to_rows(reference_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                new_reference_sheet.cell(row=r_idx, column=c_idx, value=value)

        forecast_sheet.freeze_panes = 'A2'
        new_reference_sheet.freeze_panes = 'A2'

        logging.info("Adjusting column widths and applying styles")
        for sheet in [forecast_sheet, new_reference_sheet]:
            style_worksheet(sheet)

        workbook.remove(reference_sheet)
        new_reference_sheet.title = "Reference"

        if 'Sheet1' in workbook.sheetnames:
            std = workbook['Sheet1']
            workbook.remove(std)

        set_forecast_sheet_as_active(workbook)

        logging.info(f"Saving workbook to {output_filepath}")
        workbook.save(output_filepath)
        logging.info(f"Data saved to {output_filepath}")

    except Exception as e:
        logging.error(f"An error occurred: {e}")
        # show_message("Error", f"An error occurred: {e}", type='error')


def set_forecast_sheet_as_active(workbook):
    if "Working" in workbook.sheetnames:
        workbook.active = workbook.sheetnames.index("Working")


def main(args):
    file_path = args.get('file_path')
    if not file_path or not os.path.exists(file_path):
        logging.error(f"The specified file does not exist: {file_path}")
        return

    references_month = int(args.get('references_month', 6))
    references_year = int(args.get('references_year', 2024))
    target_start_year = int(args.get('target_start_year', 2025))
    target_end_year = int(args.get('target_end_year', 2025))
    specifics_enabled = args.get('specifics_enabled', False)
    prod_nums = args.get('prod_nums', [])
    bus_chanl_nums = args.get('bus_chanl_nums', [])
    output_dir = args.get('output_dir')
    if not output_dir or not os.path.exists(output_dir):
        logging.error(f"The specified output directory does not exist: {output_dir}")
        return

    df = load_excel(file_path)

    forecast_df, reference_df = calculate_forecast(df, references_month, references_year, target_start_year, target_end_year, specifics_enabled, prod_nums, bus_chanl_nums)
    if not forecast_df.empty:
        save_dataframe_with_formatting(forecast_df, reference_df, output_dir, file_path, references_year, prod_nums, bus_chanl_nums)

if __name__ == "__main__":
    if len(sys.argv) > 1:
        args = json.loads(sys.argv[1])
    else:
        args = {}
    main(args)