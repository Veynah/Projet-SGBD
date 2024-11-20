import os
import sys
import warnings
import re
import logging
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from collections import defaultdict

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REFERENCE_DATA_DIR = os.path.join(BASE_DIR, 'inputs')
OUTPUT_DIR = os.path.join(BASE_DIR, 'outputs/tsv')

def ensure_output_dir():
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    if not os.path.exists(REFERENCE_DATA_DIR):
        os.makedirs(REFERENCE_DATA_DIR)


def extract_text_from_xlsx(xlsx_path):
    try:
        workbook = load_workbook(filename=xlsx_path, data_only=True)
    except PermissionError as e:
        logging.error(f"Permission denied: {e}. Skipping file {xlsx_path}")
        return None
    except Exception as e:
        logging.error(f"Failed to load workbook: {e}")
        return None

    text_content = []
    channels_included = []
    exclusion_list = [
        "REQUEST FOR PO BROADCASTING CONTENT", "vendor data", "!! COMPANY ISSUING INVOICES !!", "(landcode + number)",
        "(if existing in sap)", "NAME OF CHANNEL", "COMFORT / BOUQUET…", "new", "old",
        "(if fee varies per range of number of subcribers, pls indicate in detail underneath)",
        "from … to …. number of subscribers", "(describe pls)", "(explain briefly)",
        "(invoice dated beginning of invoicing period or end of invoicing period)",
        "CALCULATION OF INDEX",
        "berekend op het aantal abonnees op het einde van elk kwartaal in het verzorgingsgebied van de omroeporganisatie",
        "(cd remarks)"
    ]

    in_channels_section = False
    stop_processing = False

    period = extract_period_from_filename(xlsx_path)
    if period:
        text_content.append("CONTRACT PERIOD")
        text_content.append(period)

    for sheet in workbook:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if stop_processing:
                    break
                column = cell.column_letter if isinstance(cell, Cell) else cell.coordinate.split(str(cell.row))[0]
                if cell.value is not None and cell.data_type != 'f':
                    cell_value_str = str(cell.value).strip()

                    if any(exclusion_phrase.lower() in cell_value_str.lower() for exclusion_phrase in exclusion_list):
                        continue

                    if "ADDITIONAL INFORMATION" in cell_value_str or "ADDITIONAL INFO" in cell_value_str:
                        stop_processing = True
                        break

                    if "CHANNEL INFORMATION" in cell_value_str:
                        in_channels_section = True
                        continue

                    if "DELIVERY PERIOD/DATE" in cell_value_str:
                        in_channels_section = False
                        text_content.append(cell_value_str)
                        continue

                    if in_channels_section:
                        channels_included.append(cell_value_str)
                    else:
                        text_content.append(cell_value_str)

    if channels_included:
        channels_text = parse_channel_information(channels_included)
        text_content.append(channels_text)

    return "\n".join(text_content)


def parse_channel_information(channels_included):
    existing_packs = load_existing_packs()

    existing_channels = [
        "A", "AB3", "ABXplore", "Action", "Al Jazeera English", "Animal Planet", "Animal Planet SD NL",
        "Animal Planet SD FR", "Animaux",
        "Antenne Centre Télévision", "Automoto", "Arte", "B", "Baby TV", "BBC First",
        "BBC News", "Be 1", "BRF TV", "C",
        "Cartoon Network", "Cartoonito",
        "CGTN", "China Global Television Network", "CNBC Europe", "CRIME DISTRICT",
        "CNN International", "Comedy Central", "Crime district", "D",
        "Discovery Channel", "Discovery Channel SD NL", "Discovery Channel SD FR", "Discovery Science",
        "Disney Channel", "Disney Channel", "Discovery Channel HD FR", "Discovery Channel HD NL",
        "Discovery World SD NL", "Discovery World SD FR", "E",
        "E!", "VRT 1", "ESPN Classic",
        "Euronews", "Eurosport", "Eurosport 1", "Eurosport 2", "EUX.TV", "F",
        "Fox Life", "H", "History", "I",
        "Investigation Discovery", "Investigations Discovery NL", "Investigations Discovery FR", "J", "JIM", "K",
        "Kadet", "Ketnet", "M", "M6 Boutique", "Mangas", "MTV", "N",
        "National Geographic", "National Geographic Wild",
        "NickMusic EMEA", "Nick Jr.", "Nickelodeon",
        "Nickelodeon", "Nicktoons", "P", "Pebble TV", "Play More", "Play4",
        "Play5", "Play6", "Prime Action", "Prime Family", "Prime Fezztival",
        "Prime Series", "Prime Star", "Private Spice", "Q", "Qmusic TV", "R", "Regionale Televiesieomroep TV Limburg",
        "RT", "RTBF", "RTL Club", "RTL Plug", "RTL-TVI", "S", "Science et Vie TV", "ShortsTV",
        "Star Channel", "Stingray Classica", "Stingray Djazz",
        "Stingray iConcerts", "Stingray Lite TV", "T", "Tipik", "TiVi5 Monde", "TLC HD NL",
        "TMF Dance", "TMF NL", "TMF Pure", "TNT", "Trek", "La Trois",
        "TV Oranje", "TV5Monde", "U", "La Une", "V", "VRT Canvas", "VTM",
        "VTM 2", "VTM 3", "VTM Kids", "VTM Non-Stop Dokters", "X", "Xite"
    ]

    possible_packs_set = set(pack.lower() for pack in existing_packs)
    possible_channels_set = set(channel.lower() for channel in existing_channels)
    channels = defaultdict(set)
    current_channel = None

    for line in channels_included:
        line = line.strip().lower()
        if not line:
            continue

        if line in possible_packs_set:
            if current_channel:
                channels[current_channel].add(line)
        elif line in possible_channels_set:
            current_channel = line
        elif "&" in line:
            parts = line.split("&")
            full_pack_name = line
            if full_pack_name in possible_packs_set and current_channel:
                channels[current_channel].add(full_pack_name)
            else:
                for part in parts:
                    part = part.strip()
                    if current_channel and part in possible_packs_set:
                        channels[current_channel].add(part)

    output_lines = []
    for channel, packs in channels.items():
        packs_str = ", ".join(sorted(packs))
        output_lines.append(f"{channel.capitalize()} ({packs_str})")

    return "\n".join(output_lines)


def extract_period_from_filename(filename):
    patterns = [
        r'\b(\d{4}-\d{4})\b',  # range of 4-digit years
        r'\b(\d{4}-\d{2})\b',  # 4 digit year - 2 digit year ie: 2024-25
        r'\b(\d{4})\b',  # 4 digit year
        r'\b(\d{2})\b'  #2 digit year
    ]

    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            period = match.group(1)
            if '-' in period:
                parts = period.split('-')
                if len(parts[1]) == 2:
                    period = parts[0] + '-' + parts[0][:2] + parts[1]
                return period
            elif len(period) == 2:
                period = '20' + period
            return period
    return None


def load_existing_packs():
    latest_files = glob.glob(os.path.join(REFERENCE_DATA_DIR, 'Product_Grouping_Latest*.xlsx'))
    if latest_files:
        latest_file = max(latest_files, key=os.path.getctime)
        logging.info(f"Using latest Product_Grouping_Latest file: {latest_file}")

        df = pd.read_excel(latest_file, usecols=['PROD_MSY_GRP'])
        return df['PROD_MSY_GRP'].tolist()

    packs_tv_file = os.path.join(REFERENCE_DATA_DIR, 'packsTV.tsv')
    if os.path.exists(packs_tv_file):
        logging.info(f"Using packsTV.tsv file: {packs_tv_file}")
        packs_tv_df = pd.read_csv(packs_tv_file, sep='\t')
        return packs_tv_df['PROD_MSY_GRP'].tolist()

    logging.error("No suitable reference file found.")
    sys.exit(1)


def process_directory(folder_path):
    if not os.path.exists(folder_path):
        logging.error(f"The directory {folder_path} does not exist.")
        sys.exit(1)

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            logging.info(f"Processing now: {filename}")
            xlsx_path = os.path.join(folder_path, filename)
            text_content = extract_text_from_xlsx(xlsx_path)

            if text_content:
                tsv_filename = os.path.splitext(filename)[0] + ".tsv"
                tsv_path = os.path.join(OUTPUT_DIR, tsv_filename)

                with open(tsv_path, "w", newline='', encoding='utf-8') as tsv_file:
                    tsv_file.write(text_content)
                logging.info(f"DONE: {filename}, exported in {tsv_filename}")

    latest_files = glob.glob(os.path.join(REFERENCE_DATA_DIR, 'Product_Grouping_Latest*.xlsx'))
    if not latest_files:
        logging.error("No files starting with 'Product_Grouping_Latest' found in 'inputs' directory.")
        return

    latest_file = max(latest_files, key=os.path.getctime)
    logging.info(f"Latest file found: {latest_file}")

    df = pd.read_excel(latest_file, usecols=['PROD_MSY_GRP'])

    packs_tv_file = os.path.join(REFERENCE_DATA_DIR, 'packsTV.tsv')
    if not os.path.exists(packs_tv_file):
        logging.error(f"{packs_tv_file} does not exist.")
        return

    packs_tv_df = pd.read_csv(packs_tv_file, sep='\t')
    existing_packs = packs_tv_df['PROD_MSY_GRP'].tolist()
    logging.info(f"Loaded {len(existing_packs)} packs from {packs_tv_file}")

    additional_df = pd.DataFrame(existing_packs, columns=['PROD_MSY_GRP'])

    combined_df = pd.concat([df, additional_df])

    df_unique = combined_df.drop_duplicates()

    tsv_file = os.path.join(OUTPUT_DIR, 'PROD_MSY_GRP_unique.tsv')
    df_unique.to_csv(tsv_file, sep='\t', index=False)
    logging.info(f"Column 'PROD_MSY_GRP' with unique values has been exported to {tsv_file}")


def save_to_tsv(content, file_path):
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    output_path = os.path.join(OUTPUT_DIR, f"{base_name}.tsv")

    if not output_path.endswith('.tsv'):
        logging.error(f"Attempted to save a file with a non-TSV extension: {output_path}")
        return

    try:
        with open(output_path, 'w') as file:
            file.write(content)
        logging.info(f"Successfully saved: {output_path}")
    except Exception as e:
        logging.error(f"Error saving file {output_path}: {e}")




def main():
    if len(sys.argv) < 2:
        logging.error("No file path provided")
        return

    file_path = sys.argv[1]
    if not os.path.isfile(file_path):
        logging.error(f"File does not exist: {file_path}")
        return

    ensure_output_dir()

    result = extract_text_from_xlsx(file_path)
    if result:
        print(result)
        save_to_tsv(result, file_path)
    else:
        logging.error("Failed to extract text")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    main()



